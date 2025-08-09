import os
import logging
import mimetypes
from typing import Tuple, Optional, Dict, Any
import tempfile
import json
from werkzeug.utils import secure_filename
import uuid
from datetime import datetime

# 导入文档处理库
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("PyPDF2未安装，PDF处理功能不可用")

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False
    logging.warning("python-docx未安装，Word文档处理功能不可用")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl未安装，Excel处理功能不可用")

try:
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False
    logging.warning("PIL或pytesseract未安装，图片OCR功能不可用")

from .r2_storage import R2Storage

logger = logging.getLogger(__name__)

class AttachmentProcessor:
    """附件处理器 - 负责文件上传、存储和内容提取"""
    
    # 支持的文件类型
    SUPPORTED_FILE_TYPES = {
        'text': ['.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml', '.csv'],
        'document': ['.pdf', '.docx', '.doc'],
        'spreadsheet': ['.xlsx', '.xls'],
        'image': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff'],
        'archive': ['.zip', '.rar', '.7z'],
        'code': ['.py', '.js', '.java', '.cpp', '.c', '.php', '.rb', '.go', '.rs']
    }
    
    # 文件大小限制 (50MB)
    MAX_FILE_SIZE = 50 * 1024 * 1024
    
    def __init__(self):
        """初始化附件处理器"""
        self.r2_storage = R2Storage()
        self.temp_dir = tempfile.gettempdir()
    
    def is_file_supported(self, filename: str) -> bool:
        """检查文件类型是否被支持"""
        file_ext = os.path.splitext(filename.lower())[1]
        
        # 检查基本文本和代码文件
        if file_ext in self.SUPPORTED_FILE_TYPES['text'] or file_ext in self.SUPPORTED_FILE_TYPES['code']:
            return True
            
        # 检查PDF文件（需要PyPDF2）
        if file_ext in self.SUPPORTED_FILE_TYPES['document']:
            if file_ext == '.pdf':
                return PDF_AVAILABLE
            elif file_ext in ['.docx', '.doc']:
                return DOCX_AVAILABLE
            
        # 检查Excel文件（需要openpyxl）
        if file_ext in self.SUPPORTED_FILE_TYPES['spreadsheet']:
            return EXCEL_AVAILABLE
            
        # 检查图片文件（需要PIL和pytesseract）
        if file_ext in self.SUPPORTED_FILE_TYPES['image']:
            return OCR_AVAILABLE
            
        # 其他类型暂不支持
        return False
    
    def get_file_type(self, filename: str) -> str:
        """获取文件类型分类"""
        file_ext = os.path.splitext(filename.lower())[1]
        for type_name, extensions in self.SUPPORTED_FILE_TYPES.items():
            if file_ext in extensions:
                return type_name
        return 'unknown'
    
    def validate_file(self, file) -> Tuple[bool, str]:
        """验证上传的文件"""
        if not file:
            return False, "没有选择文件"
        
        if file.filename == '':
            return False, "文件名为空"
        
        if not self.is_file_supported(file.filename):
            return False, f"不支持的文件类型: {file.filename}"
        
        # 检查文件大小（如果可能）
        file.seek(0, 2)  # 移动到文件末尾
        file_size = file.tell()
        file.seek(0)  # 重置文件指针
        
        if file_size > self.MAX_FILE_SIZE:
            return False, f"文件太大，最大支持 {self.MAX_FILE_SIZE // (1024*1024)}MB"
        
        return True, ""
    
    def save_temp_file(self, file) -> Tuple[bool, str]:
        """保存文件到临时目录"""
        try:
            filename = secure_filename(file.filename)
            # 生成唯一的临时文件名
            temp_filename = f"{uuid.uuid4().hex}_{filename}"
            temp_path = os.path.join(self.temp_dir, temp_filename)
            
            file.save(temp_path)
            logger.info(f"临时文件保存成功: {temp_path}")
            return True, temp_path
            
        except Exception as e:
            error_msg = f"保存临时文件失败: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def extract_text_content(self, file_path: str, file_type: str) -> Tuple[bool, str]:
        """从文件中提取文本内容"""
        try:
            content = ""
            
            if file_type == 'text' or file_type == 'code':
                # 处理文本文件
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            
            elif file_type == 'document':
                # 处理文档文件
                file_ext = os.path.splitext(file_path.lower())[1]
                
                if file_ext == '.pdf':
                    content = self._extract_pdf_text(file_path)
                elif file_ext in ['.docx', '.doc']:
                    content = self._extract_docx_text(file_path)
            
            elif file_type == 'spreadsheet':
                # 处理电子表格
                content = self._extract_excel_text(file_path)
            
            elif file_type == 'image':
                # 处理图片 - OCR文字识别
                content = self._extract_image_text(file_path)
            
            else:
                return False, f"不支持提取该文件类型的内容: {file_type}"
            
            if not content.strip():
                return False, "未能从文件中提取到有效内容"
            
            return True, content
            
        except Exception as e:
            error_msg = f"提取文件内容失败: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """从PDF文件提取文本"""
        if not PDF_AVAILABLE:
            return "PDF处理功能不可用，请安装PyPDF2库"
        
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
                return text
        except Exception as e:
            logger.error(f"PDF文本提取失败: {str(e)}")
            return ""
    
    def _extract_docx_text(self, file_path: str) -> str:
        """从DOCX文件提取文本"""
        if not DOCX_AVAILABLE:
            return "Word文档处理功能不可用，请安装python-docx库"
        
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"DOCX文本提取失败: {str(e)}")
            return ""
    
    def _extract_excel_text(self, file_path: str) -> str:
        """从Excel文件提取文本"""
        if not EXCEL_AVAILABLE:
            return "Excel处理功能不可用，请安装openpyxl库"
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"工作表: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
        except Exception as e:
            logger.error(f"Excel文本提取失败: {str(e)}")
            return ""
    
    def _extract_image_text(self, file_path: str) -> str:
        """从图片中提取文字 (OCR)"""
        if not OCR_AVAILABLE:
            return "图片OCR功能不可用，请安装PIL和pytesseract库"
        
        try:
            # 需要安装tesseract
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            return text
        except Exception as e:
            logger.error(f"图片文字提取失败: {str(e)}")
            return "图片文字识别功能需要安装tesseract-ocr"
    
    def upload_to_r2(self, file_path: str, original_filename: str) -> Tuple[bool, str]:
        """上传文件到Cloudflare R2"""
        try:
            # 生成R2存储路径
            file_ext = os.path.splitext(original_filename)[1]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            object_key = f"attachments/{timestamp}_{uuid.uuid4().hex}{file_ext}"
            
            # 设置Content-Type
            content_type, _ = mimetypes.guess_type(original_filename)
            
            success, result = self.r2_storage.upload_file(
                file_path, 
                object_key, 
                content_type=content_type
            )
            
            if success:
                return True, object_key
            else:
                return False, result
                
        except Exception as e:
            error_msg = f"上传到R2失败: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def process_attachment(self, file, user_id: str) -> Dict[str, Any]:
        """处理附件的完整流程"""
        result = {
            'success': False,
            'message': '',
            'attachment_info': None
        }
        
        temp_file_path = None
        
        try:
            # 1. 验证文件
            valid, error_msg = self.validate_file(file)
            if not valid:
                result['message'] = error_msg
                return result
            
            # 2. 保存临时文件
            success, temp_file_path = self.save_temp_file(file)
            if not success:
                result['message'] = temp_file_path
                return result
            
            # 3. 确定文件类型
            file_type = self.get_file_type(file.filename)
            
            # 4. 提取文件内容
            success, content = self.extract_text_content(temp_file_path, file_type)
            if not success:
                logger.warning(f"内容提取失败: {content}")
                content = ""  # 继续处理，但没有文本内容
            
            # 5. 上传到R2
            success, r2_key = self.upload_to_r2(temp_file_path, file.filename)
            if not success:
                result['message'] = r2_key
                return result
            
            # 6. 构建附件信息
            attachment_info = {
                'id': str(uuid.uuid4()),
                'original_filename': file.filename,
                'file_type': file_type,
                'file_size': os.path.getsize(temp_file_path),
                'r2_key': r2_key,
                'upload_time': datetime.now().isoformat(),
                'user_id': user_id,
                'content_preview': content[:500] if content else "",  # 前500字符预览
                'has_text_content': bool(content)
            }
            
            result['success'] = True
            result['message'] = '附件上传并处理成功'
            result['attachment_info'] = attachment_info
            result['full_content'] = content  # 完整内容用于AI分析
            
        except Exception as e:
            error_msg = f"附件处理失败: {str(e)}"
            logger.error(error_msg)
            result['message'] = error_msg
            
        finally:
            # 清理临时文件
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                    logger.info(f"临时文件已清理: {temp_file_path}")
                except Exception as e:
                    logger.warning(f"清理临时文件失败: {str(e)}")
        
        return result
    
    def get_attachment_download_url(self, r2_key: str) -> Tuple[bool, str]:
        """获取附件的下载URL"""
        return self.r2_storage.get_file_url(r2_key, expires_in=3600)
